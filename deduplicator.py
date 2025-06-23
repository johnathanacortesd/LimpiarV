# deduplicator.py

# --- INICIO DE LAS IMPORTACIONES ---
import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
from difflib import SequenceMatcher
from collections import defaultdict
import re
import datetime
from copy import deepcopy
import html

# --- CONSTANTES Y FUNCIONES AUXILIARES ---
CONSERVAR = "Conservar"
ELIMINAR = "Eliminar"
SI = "Sí"
NO = "FALSE"
TONO_DUPLICADA = "Duplicada"
TEMA_VACIO = "-"

def norm_key(text):
    return re.sub(r'\W+', '', str(text).lower().strip()) if text else ""

def convert_html_entities(text):
    if not isinstance(text, str): return text
    return html.unescape(text)

def normalize_title(title):
    if not isinstance(title, str):
        return ""
    processed_title = convert_html_entities(title)
    processed_title = processed_title.lower()
    processed_title = processed_title.strip()
    processed_title = re.sub(r'\s*\|\s*[\w\s]+$', '', processed_title)
    processed_title = re.sub(r'\W+', '', processed_title)
    return processed_title

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = convert_html_entities(text)
    text = text.replace('<br>', ' ').replace('[...]', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    if match := re.search(r'[A-Z]', text):
        text = text[match.start():]
    if text and not text.endswith('...'):
        text = re.sub(r'[\.,;:]$', '', text.strip()).strip()
        text += '...'
    return text

def extract_link(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return {"value": cell.value or "Link", "url": cell.hyperlink.target}
    if cell.value and isinstance(cell.value, str):
        if match := re.search(r'=HYPERLINK\("([^"]+)"', cell.value):
            return {"value": "Link", "url": match.group(1)}
    return {"value": cell.value, "url": None}

def format_date(date_val):
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, str):
        try:
            return datetime.datetime.strptime(date_val.split(' ')[0], '%Y-%m-%d').strftime('%Y-%m-%d')
        except ValueError:
            return str(date_val)
    return str(date_val)

def parse_date_obj(date_val):
    if isinstance(date_val, (datetime.datetime, datetime.date)): return date_val
    if isinstance(date_val, str):
        try:
            if '/' in date_val:
                return datetime.datetime.strptime(date_val.split(' ')[0], '%d/%m/%Y')
            else:
                return datetime.datetime.strptime(date_val.split(' ')[0], '%Y-%m-%d')
        except (ValueError, AttributeError):
            return datetime.datetime.min
    return datetime.datetime.min

def parse_time_obj(time_val):
    if isinstance(time_val, datetime.time): return time_val
    if isinstance(time_val, datetime.datetime): return time_val.time()
    if isinstance(time_val, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try: return datetime.datetime.strptime(time_val, fmt).time()
            except ValueError: continue
    return datetime.time.min

def es_internet(row):
    return norm_key(row.get(norm_key('Tipo de Medio'))) == 'internet'

def is_title_problematic(title):
    if not isinstance(title, str) or not title.strip(): return True
    if re.search(r'^(sin t[íi]tulo|untitled|no title)$', title.lower()): return True
    if re.search(r'[Ââ€™"""'']', title): return True
    if re.search(r'\s*\|\s*[\w\s]+$', title): return True
    return False

def mark_as_duplicate_to_delete(row):
    row['Mantener'] = ELIMINAR
    row[norm_key('Tono')] = TONO_DUPLICADA
    row[norm_key('Tema')] = TEMA_VACIO
    row[norm_key('Temas Generales - Tema')] = TEMA_VACIO

def get_title_priority(row):
    medio_norm = norm_key(row.get(norm_key('Medio')))
    titulo_str = str(row.get(norm_key('Título'), ''))
    if medio_norm == norm_key('El Colombiano (Online)'): return 1 if '| El Colombiano' in titulo_str else 0
    if medio_norm == norm_key('El Nuevo Siglo (Online)'): return 1 if titulo_str.strip().endswith('El Nuevo Siglo') else 0
    return 0

def get_title_cleanliness_score(row):
    original_title = str(row.get('original_titulo', ''))
    cleaned_title = str(row.get(norm_key('Título'), ''))
    return 0 if original_title == cleaned_title else 1

def are_rows_similar(row_i, row_j):
    SIMILARIDAD_MINIMA = 0.85
    if norm_key(row_i.get(norm_key('Medio'))) != norm_key(row_j.get(norm_key('Medio'))):
        return False
    if norm_key(row_i.get(norm_key('Menciones - Empresa'))) != norm_key(row_j.get(norm_key('Menciones - Empresa'))):
        return False
    date_i = parse_date_obj(row_i.get(norm_key('Fecha')))
    date_j = parse_date_obj(row_j.get(norm_key('Fecha')))
    if not (date_i and date_j and abs((date_i - date_j).days) <= 1):
        return False
    title_i = normalize_title(row_i.get(norm_key('Título')))
    title_j = normalize_title(row_j.get(norm_key('Título')))
    if not (title_i and title_j and SequenceMatcher(None, title_i, title_j).ratio() >= SIMILARIDAD_MINIMA):
        return False
    return True

def run_deduplication_process(wb, internet_dict, region_dict):
    sheet = wb.active
    custom_link_style = NamedStyle(name="CustomLink", font=Font(color="0000FF", underline="single"), alignment=Alignment(horizontal="left"))
    if "CustomLink" not in wb.named_styles:
        wb.add_named_style(custom_link_style)
    headers = [cell.value for cell in sheet[1]]
    headers_norm = [norm_key(h) for h in headers]
    processed_rows = []

    for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2)):
        if all(c.value is None for c in row_cells): continue
        base_data = {'original_row_index': row_idx + 2}
        for i, cell in enumerate(row_cells):
            col_name = headers_norm[i]
            if col_name in [norm_key('Link Nota'), norm_key('Link (Streaming - Imagen)')]:
                base_data[col_name] = extract_link(cell)
            else:
                base_data[col_name] = cell.value
        titulo_key = norm_key('Título')
        original_title = str(base_data.get(titulo_key, ''))
        base_data['original_titulo'] = original_title
        base_data[titulo_key] = convert_html_entities(original_title)
        base_data[norm_key('Resumen - Aclaracion')] = corregir_texto(base_data.get(norm_key('Resumen - Aclaracion')))
        tipo_medio_key = norm_key('Tipo de Medio')
        tm_norm = norm_key(base_data.get(tipo_medio_key))
        if tm_norm in {'aire', 'cable'}: base_data[tipo_medio_key] = 'Televisión'
        elif tm_norm in {'am', 'fm'}: base_data[tipo_medio_key] = 'Radio'
        elif tm_norm == 'diario': base_data[tipo_medio_key] = 'Prensa'
        elif tm_norm == 'online': base_data[tipo_medio_key] = 'Internet'
        elif tm_norm == 'revista': base_data[tipo_medio_key] = 'Revista'
        link_nota_key, link_streaming_key = norm_key("Link Nota"), norm_key("Link (Streaming - Imagen)")
        tipo_medio_val = base_data.get(tipo_medio_key)
        if tipo_medio_val == "Internet":
            base_data[link_nota_key], base_data[link_streaming_key] = (base_data.get(link_streaming_key), base_data.get(link_nota_key))
        elif tipo_medio_val in {"Prensa", "Revista"}:
            is_link_nota_empty = not base_data.get(link_nota_key) or not base_data.get(link_nota_key, {}).get('url')
            has_streaming_link = base_data.get(link_streaming_key, {}).get('url')
            if is_link_nota_empty and has_streaming_link: base_data[link_nota_key] = base_data.get(link_streaming_key)
            base_data[link_streaming_key] = None
        elif tipo_medio_val in {"Radio", "Televisión"}: base_data[link_streaming_key] = None
        
        menciones_key = norm_key('Menciones - Empresa')
        menciones_str = str(base_data.get(menciones_key) or '')
        menciones = [m.strip() for m in menciones_str.split(';') if m.strip()]
        if not menciones:
            processed_rows.append(base_data)
        else:
            for mencion in menciones:
                new_row = deepcopy(base_data)
                new_row[menciones_key] = mencion
                processed_rows.append(new_row)

    medio_key, tipo_medio_key, region_key = norm_key('Medio'), norm_key('Tipo de Medio'), norm_key('Región')
    for row in processed_rows:
        if str(row.get(tipo_medio_key, '')).lower().strip() == 'internet':
            medio_val = str(row.get(medio_key, '')).lower().strip()
            if medio_val in internet_dict: row[medio_key] = internet_dict[medio_val]
        medio_actual_val = str(row.get(medio_key, '')).lower().strip()
        row[region_key] = region_dict.get(medio_actual_val, "Online")

    for row in processed_rows:
        row.update({'Duplicada': NO, 'Posible Duplicada': NO, 'Mantener': CONSERVAR, 'ID Fila Conservada': ''})
    
    for row in processed_rows:
        if is_title_problematic(row.get(norm_key('Título'))):
            row['Mantener'] = ELIMINAR
            row['Duplicada'] = SI
            
    grupos_exactos = defaultdict(list)
    for idx, row in enumerate(processed_rows):
        if row['Mantener'] == ELIMINAR: continue
        key_tuple = (normalize_title(row.get(norm_key('Título'))), norm_key(row.get(norm_key('Medio'))), norm_key(row.get(norm_key('Menciones - Empresa'))), format_date(row.get(norm_key('Fecha'))))
        if not es_internet(row): key_tuple += (str(row.get(norm_key('Hora'))),)
        grupos_exactos[key_tuple].append(idx)
        
    for indices in grupos_exactos.values():
        if len(indices) > 1:
            indices.sort(key=lambda i: get_title_cleanliness_score(processed_rows[i]))
            indices.sort(key=lambda i: get_title_priority(processed_rows[i]), reverse=True)
            indices.sort(key=lambda i: "'" in processed_rows[i]['original_titulo'] or '"' in processed_rows[i]['original_titulo'], reverse=True)
            indices.sort(key=lambda i: (parse_date_obj(processed_rows[i].get(norm_key('Fecha'))), parse_time_obj(processed_rows[i].get(norm_key('Hora')))), reverse=True)
            winner_idx = indices[0]
            winner_id = str(processed_rows[winner_idx].get(norm_key('ID Noticia'), ''))
            processed_rows[winner_idx]['Duplicada'] = SI
            for loser_idx in indices[1:]:
                mark_as_duplicate_to_delete(processed_rows[loser_idx])
                processed_rows[loser_idx]['Duplicada'] = SI
                processed_rows[loser_idx]['ID Fila Conservada'] = winner_id

    final_clusters = []
    candidates_indices = []
    for idx, row in enumerate(processed_rows):
        if row['Duplicada'] == NO and row['Mantener'] == CONSERVAR:
            candidates_indices.append(idx)
    for idx in candidates_indices:
        row_i = processed_rows[idx]
        found_cluster = False
        for cluster in final_clusters:
            representative_idx = cluster[0]
            row_j = processed_rows[representative_idx]
            if es_internet(row_i) != es_internet(row_j): continue
            if are_rows_similar(row_i, row_j):
                cluster.append(idx)
                found_cluster = True
                break
        if not found_cluster:
            final_clusters.append([idx])
            
    for cluster in final_clusters:
        if len(cluster) > 1:
            cluster.sort(key=lambda i: get_title_cleanliness_score(processed_rows[i]))
            cluster.sort(key=lambda i: get_title_priority(processed_rows[i]), reverse=True)
            cluster.sort(key=lambda i: "'" in processed_rows[i]['original_titulo'] or '"' in processed_rows[i]['original_titulo'], reverse=True)
            cluster.sort(key=lambda i: (parse_date_obj(processed_rows[i].get(norm_key('Fecha'))), parse_time_obj(processed_rows[i].get(norm_key('Hora')))), reverse=True)
            winner_idx = cluster[0]
            winner_id = str(processed_rows[winner_idx].get(norm_key('ID Noticia'), ''))
            processed_rows[winner_idx]['Posible Duplicada'] = SI
            for loser_idx in cluster[1:]:
                if processed_rows[loser_idx]['Mantener'] != ELIMINAR:
                    mark_as_duplicate_to_delete(processed_rows[loser_idx])
                processed_rows[loser_idx]['Posible Duplicada'] = SI
                processed_rows[loser_idx]['ID Fila Conservada'] = winner_id
        
    final_order = ["ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Sección - Programa", "Región", "Título", "Autor - Conductor", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia", "Tono", "Tema", "Temas Generales - Tema", "Resumen - Aclaracion", "Link Nota", "Link (Streaming - Imagen)", "Menciones - Empresa", "Duplicada", "Posible Duplicada", "Mantener", "ID Fila Conservada"]
    main_wb = openpyxl.Workbook()
    main_sheet = main_wb.active
    main_sheet.title = "Resultado"
    main_sheet.append(final_order)
    if "CustomLink" not in main_wb.named_styles: main_wb.add_named_style(custom_link_style)
    nissan_wb = openpyxl.Workbook()
    nissan_sheet = nissan_wb.active
    nissan_sheet.title = "Resumen Concatenado"
    nissan_sheet.append(["Resumen"])
    processed_rows.sort(key=lambda r: r.get('original_row_index', 0))
    
    mantener_key = norm_key("Mantener")
    titulo_key = norm_key("Título")
    resumen_key = norm_key("Resumen - Aclaracion")
    
    for row_data in processed_rows:
        if row_data.get(mantener_key, row_data.get('Mantener')) == CONSERVAR:
            title = str(row_data.get(titulo_key, ''))
            title = re.sub(r'\s*\|\s*[\w\s]+$', '', title).strip()
            row_data[titulo_key] = title
            resumen_aclaracion = str(row_data.get(resumen_key, ''))
            concatenated_summary = f"{title} {resumen_aclaracion}".strip()
            nissan_sheet.append([concatenated_summary])
        
        new_row_to_append = []
        for header in final_order:
            val = row_data.get(norm_key(header), row_data.get(header, None))
            new_row_to_append.append(val.get('value') if isinstance(val, dict) else val)
        main_sheet.append(new_row_to_append)
        
    link_nota_idx = final_order.index("Link Nota")
    link_streaming_idx = final_order.index("Link (Streaming - Imagen)")
    
    # --- INICIO DE LA SECCIÓN CORREGIDA ---
    # La lógica anterior con `row_map` era la causa del error.
    # Esta nueva lógica es más simple y segura, ya que el orden de `processed_rows`
    # y las filas en `main_sheet` es idéntico.
    for i, sheet_row_cells in enumerate(main_sheet.iter_rows(min_row=2)):
        if i < len(processed_rows):
            processed = processed_rows[i]
            
            link_data = processed.get(norm_key("Link Nota"))
            if isinstance(link_data, dict) and link_data.get("url"):
                cell = sheet_row_cells[link_nota_idx]
                cell.hyperlink = link_data["url"]
                cell.value = "Link"
                cell.style = "CustomLink"
            
            link_data_stream = processed.get(norm_key("Link (Streaming - Imagen)"))
            if isinstance(link_data_stream, dict) and link_data_stream.get("url"):
                cell = sheet_row_cells[link_streaming_idx]
                cell.hyperlink = link_data_stream["url"]
                cell.value = "Link"
                cell.style = "CustomLink"
    # --- FIN DE LA SECCIÓN CORREGIDA ---

    summary = {
        "total_rows": len(processed_rows),
        "to_eliminate": sum(1 for r in processed_rows if r.get(mantener_key, r.get('Mantener')) == ELIMINAR),
        "to_conserve": len(processed_rows) - sum(1 for r in processed_rows if r.get(mantener_key, r.get('Mantener')) == ELIMINAR),
        "exact_duplicates": sum(1 for r in processed_rows if r.get('Duplicada') == SI),
        "possible_duplicates": sum(1 for r in processed_rows if r.get('Posible Duplicada') == SI and r.get('Duplicada') == NO)
    }
    return main_wb, nissan_wb, summary
