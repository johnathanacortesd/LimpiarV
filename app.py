import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import datetime
import io
import re
import html
import numpy as np
import pyarrow as pa

# --- Configuración de la página ---
st.set_page_config(page_title="Procesador de Dossiers (Lite) v1.8", layout="wide")

# ==============================================================================
# SECCIÓN DE FUNCIONES AUXILIARES
# ==============================================================================

def extract_link_from_cell(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None

def convert_html_entities(text):
    """
    Convierte entidades HTML mal codificadas a caracteres normales.
    Maneja tanto entidades hexadecimales como decimales.
    """
    if not isinstance(text, str):
        return text
    
    # Primero decodificar entidades HTML estándar
    text = html.unescape(text)

    # Manejar entidades HTML numéricas hexadecimales específicas
    html_entities = {
        '&#xF3;': 'ó',  # ó
        '&#xE1;': 'á',  # á
        '&#xE9;': 'é',  # é
        '&#xED;': 'í',  # í
        '&#xFA;': 'ú',  # ú
        '&#xF1;': 'ñ',  # ñ
        '&#xDC;': 'Ü',  # Ü
        '&#xFC;': 'ü',  # ü
        '&#xC1;': 'Á',  # Á
        '&#xC9;': 'É',  # É
        '&#xCD;': 'Í',  # Í
        '&#xD3;': 'Ó',  # Ó
        '&#xDA;': 'Ú',  # Ú
        '&#xD1;': 'Ñ',  # Ñ
        '&#xC7;': 'Ç',  # Ç
        '&#xE7;': 'ç',  # ç
    }

    # Reemplazar entidades HTML numéricas hexadecimales
    for entity, char in html_entities.items():
        text = text.replace(entity, char)

    # Patrón para capturar otras entidades hexadecimales que no estén en el diccionario
    def replace_hex_entity(match):
        try:
            hex_code = match.group(1)
            char_code = int(hex_code, 16)
            return chr(char_code)
        except (ValueError, OverflowError):
            return match.group(0)  # Devolver original si no se puede convertir

    text = re.sub(r'&#x([0-9A-Fa-f]+);', replace_hex_entity, text)

    # Patrón para entidades decimales
    def replace_decimal_entity(match):
        try:
            decimal_code = int(match.group(1))
            return chr(decimal_code)
        except (ValueError, OverflowError):
            return match.group(0)  # Devolver original si no se puede convertir

    text = re.sub(r'&#(\d+);', replace_decimal_entity, text)

    # Limpiar caracteres problemáticos adicionales
    custom_replacements = {
        '"': '"',
        '"': '"', 
        ''': "'", 
        ''': "'", 
        'Â': '', 
        'â': '', 
        '€': '', 
        '™': ''
    }

    for entity, char in custom_replacements.items():
        text = text.replace(entity, char)

    return text

def normalize_title_for_comparison(title):
    """
    Normaliza el título para comparación de duplicados.
    """
    if not isinstance(title, str):
        return ""
    
    # Limpiar entidades HTML primero
    title = convert_html_entities(title)

    # Normalizar para comparación (remover caracteres especiales y convertir a minúsculas)
    return re.sub(r'\W+', ' ', title).lower().strip()

def clean_title_for_output(title):
    """
    ÚNICAMENTE limpia entidades HTML mal codificadas.
    NO corta, NO modifica, NO remueve ninguna parte del título.
    Solo convierte caracteres como ó a ó
    """
    if not isinstance(title, str):
        return ""
    
    # SOLO limpiar entidades HTML - NO tocar nada más
    title = convert_html_entities(title)

    # Solo quitar espacios al inicio y final, NO espacios múltiples internos
    title = title.strip()

    return title

def corregir_texto(text):
    if not isinstance(text, str): 
        return text
    text = convert_html_entities(text)
    text = re.sub(r'(<br>|[...]|\s+)', ' ', text).strip()
    match = re.search(r'[A-Z]', text)
    if match: 
        text = text[match.start():]
    if text and not text.endswith('...'): 
        text = text.rstrip('.') + '...'
    return text

def to_excel_from_df(df, final_order):
    """
    Genera archivo Excel usando openpyxl para evitar la limitación de 64k enlaces de xlsxwriter.
    Esta versión no tiene límite en la cantidad de hipervínculos.
    """
    output = io.BytesIO()
    
    # Filtrar columnas que existen en el DataFrame
    final_columns_in_df = [col for col in final_order if col in df.columns]
    df_to_excel = df[final_columns_in_df].copy()
    
    # Convertir columnas de PyArrow string a object para compatibilidad con openpyxl
    for col in df_to_excel.columns:
        if hasattr(df_to_excel[col].dtype, 'pyarrow_dtype'):
            df_to_excel[col] = df_to_excel[col].astype(object)
    
    # Crear workbook con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resultado'
    
    # Escribir encabezados con formato
    for col_idx, col_name in enumerate(df_to_excel.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
    
    # Identificar columnas de enlaces
    link_columns = set()
    for col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
        if col_name in df_to_excel.columns:
            link_columns.add(col_name)
    
    # Escribir datos fila por fila
    for row_idx, row_data in enumerate(df_to_excel.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            col_name = df_to_excel.columns[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Manejar fechas
            if col_name == 'Fecha' and pd.notna(value):
                if isinstance(value, pd.Timestamp):
                    cell.value = value.to_pydatetime()
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    cell.value = value
            # Crear hipervínculos para columnas de enlaces
            elif col_name in link_columns and pd.notna(value) and isinstance(value, str) and value.startswith('http'):
                cell.value = 'Link'
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.value = value
    
    # Ajustar anchos de columna para mejor presentación
    for col_idx, col_name in enumerate(df_to_excel.columns, start=1):
        if col_name in ['Título', 'Resumen - Aclaracion']:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 50
        elif col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
        elif col_name == 'Mantener':
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 25
        else:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    
    # Guardar en BytesIO
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

# ==============================================================================
# LÓGICA DE PROCESAMIENTO PRINCIPAL
# ==============================================================================

def run_full_process(dossier_file, config_file):
    
    st.markdown("---")
    progress_text = st.empty()

    progress_text.info("Paso 1/4: Cargando archivo de configuración...")
    try:
        config_sheets = pd.read_excel(config_file, sheet_name=None, engine='openpyxl')
        region_map = pd.Series(config_sheets['Regiones'].iloc[:, 1].values, index=config_sheets['Regiones'].iloc[:, 0].astype(str).str.lower().str.strip()).to_dict()
        internet_map = pd.Series(config_sheets['Internet'].iloc[:, 1].values, index=config_sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()).to_dict()
    except Exception as e:
        st.error(f"Error al cargar `Configuracion.xlsx`: {e}. Asegúrate de que contenga las hojas 'Regiones' e 'Internet'.")
        st.stop()

    progress_text.info("Paso 2/4: Leyendo Dossier y expandiendo filas...")
    wb = load_workbook(dossier_file)
    sheet = wb.active
    original_headers = [cell.value for cell in sheet[1] if cell.value]
    rows_to_expand = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row): 
            continue
        row_values = [c.value for c in row]
        row_data = dict(zip(original_headers, row_values))
        if 'Link Nota' in original_headers: 
            row_data['Link Nota'] = extract_link_from_cell(row[original_headers.index('Link Nota')])
        if 'Link (Streaming - Imagen)' in original_headers: 
            row_data['Link (Streaming - Imagen)'] = extract_link_from_cell(row[original_headers.index('Link (Streaming - Imagen)')])
        menciones = [m.strip() for m in str(row_data.get('Menciones - Empresa') or '').split(';') if m.strip()]
        if not menciones: 
            rows_to_expand.append(row_data)
        else:
            for mencion in menciones:
                new_row = row_data.copy()
                new_row['Menciones - Empresa'] = mencion
                rows_to_expand.append(new_row)
    df = pd.DataFrame(rows_to_expand)
    df['Estado_Duplicado'] = 'Conservar'

    progress_text.info("Paso 3/4: Aplicando limpieza, mapeos y normalizaciones...")
    
    # Optimización PyArrow: Convertir columnas de texto a string[pyarrow]
    string_columns = ['Título', 'Resumen - Aclaracion', 'Medio', 'Tipo de Medio', 
                      'Menciones - Empresa', 'Sección - Programa', 'Autor - Conductor',
                      'Región', 'Tema', 'Temas Generales - Tema', 'Tono']
    for col in string_columns:
        if col in df.columns:
            df[col] = df[col].astype('string[pyarrow]')
    
    for col in original_headers:
        if col not in df.columns: 
            df[col] = None

    # APLICAR LA LIMPIEZA CORREGIDA - SOLO ENTIDADES HTML, NO CORTAR
    df['Título'] = df['Título'].astype(str).apply(clean_title_for_output)
    df['Resumen - Aclaracion'] = df['Resumen - Aclaracion'].astype(str).apply(corregir_texto)

    tipo_medio_map = {'online': 'Internet', 'diario': 'Prensa', 'am': 'Radio', 'fm': 'Radio', 'aire': 'Televisión', 'cable': 'Televisión', 'revista': 'Revista'}
    df['Tipo de Medio'] = df['Tipo de Medio'].str.lower().str.strip().map(tipo_medio_map).fillna(df['Tipo de Medio'])
    is_internet = df['Tipo de Medio'] == 'Internet'
    is_print = df['Tipo de Medio'].isin(['Prensa', 'Revistas'])
    is_broadcast = df['Tipo de Medio'].isin(['Radio', 'Televisión'])

    df.loc[is_internet, ['Link Nota', 'Link (Streaming - Imagen)']] = df.loc[is_internet, ['Link (Streaming - Imagen)', 'Link Nota']].values
    cond_copy = is_print & df['Link Nota'].isnull() & df['Link (Streaming - Imagen)'].notnull()
    df.loc[cond_copy, 'Link Nota'] = df.loc[cond_copy, 'Link (Streaming - Imagen)']
    df.loc[is_print, 'Link (Streaming - Imagen)'] = None
    df.loc[is_broadcast, 'Link (Streaming - Imagen)'] = None

    # --- INICIO DE LA LÓGICA "CORTAR Y PEGAR" ---
    if 'Duración - Nro. Caracteres' in df.columns and 'Dimensión' in df.columns:
        # 1. Copiar el valor a la columna Dimensión para medios broadcast
        df.loc[is_broadcast, 'Dimensión'] = df.loc[is_broadcast, 'Duración - Nro. Caracteres']
        # 2. Limpiar (cortar) el valor de la columna original para esos mismos medios
        df.loc[is_broadcast, 'Duración - Nro. Caracteres'] = np.nan
    # --- FIN DE LA LÓGICA "CORTAR Y PEGAR" ---

    df['Región'] = df['Medio'].astype(str).str.lower().str.strip().map(region_map)
    df.loc[is_internet, 'Medio'] = df.loc[is_internet, 'Medio'].astype(str).str.lower().str.strip().map(internet_map).fillna(df.loc[is_internet, 'Medio'])

    progress_text.info("Paso 4/4: Detectando duplicados y generando resultados...")
    df['titulo_norm'] = df['Título'].apply(normalize_title_for_comparison)
    df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce').dt.normalize()

    df['seccion_priority'] = df['Sección - Programa'].isnull() | (df['Sección - Programa'] == '')
    df['dup_hora'] = np.where(df['Tipo de Medio'] == 'Internet', 'IGNORE_TIME', df['Hora'])

    # Crear columna Mantener inicialmente vacía
    df['Mantener'] = ''

    dup_cols_exact = ['titulo_norm', 'Medio', 'Fecha', 'Menciones - Empresa', 'dup_hora']
    sort_by_cols = dup_cols_exact + ['seccion_priority']
    ascending_order = [True] * len(dup_cols_exact) + [False]
    df.sort_values(by=sort_by_cols, ascending=ascending_order, inplace=True)
    
    # Identificar duplicados exactos y guardar el ID de la noticia a mantener
    for name, group in df.groupby(dup_cols_exact):
        if len(group) > 1:
            # El primero es el que se mantiene (tiene mejor sección)
            id_to_keep = group.iloc[0]['ID Noticia']
            # Los demás son duplicados
            duplicate_indices = group.index[1:]
            df.loc[duplicate_indices, 'Estado_Duplicado'] = 'Eliminar'
            df.loc[duplicate_indices, 'Mantener'] = f'Duplicado de: {id_to_keep}'
    
    df.sort_index(inplace=True)

    # Procesar duplicados consecutivos de Internet
    df_internet_to_check = df[(df['Estado_Duplicado'] == 'Conservar') & (is_internet)].copy()
    if not df_internet_to_check.empty:
        group_cols = ['titulo_norm', 'Medio', 'Menciones - Empresa']
        df_internet_to_check.sort_values(by=group_cols + ['Fecha'], inplace=True)
        date_diffs = df_internet_to_check.groupby(group_cols)['Fecha'].diff().dt.days
        cluster_ids = (date_diffs != 1).cumsum()
        df_internet_to_check['date_cluster'] = cluster_ids
        sort_by_cols_consecutive = group_cols + ['date_cluster', 'seccion_priority']
        ascending_order_consecutive = [True] * (len(group_cols) + 1) + [False]
        df_internet_to_check.sort_values(by=sort_by_cols_consecutive, ascending=ascending_order_consecutive, inplace=True)
        
        # Identificar duplicados consecutivos y guardar el ID de la noticia a mantener
        for name, group in df_internet_to_check.groupby(group_cols + ['date_cluster']):
            if len(group) > 1:
                id_to_keep = group.iloc[0]['ID Noticia']
                duplicate_indices = group.index[1:]
                df.loc[duplicate_indices, 'Estado_Duplicado'] = 'Eliminar'
                df.loc[duplicate_indices, 'Mantener'] = f'Duplicado de: {id_to_keep}'

    # Marcar las columnas Tono, Tema, Temas Generales como Duplicada
    df.loc[df['Estado_Duplicado'] == 'Eliminar', ['Tono', 'Tema', 'Temas Generales - Tema']] = 'Duplicada'

    st.balloons()
    progress_text.success("¡Proceso de limpieza completado! Todas las duplicadas muestran el ID de la noticia a conservar.")

    # COLUMNA MANTENER AL FINAL
    final_order = ["ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Sección - Programa", "Región", "Título", "Autor - Conductor", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia", "Tono", "Tema", "Temas Generales - Tema", "Resumen - Aclaracion", "Link Nota", "Link (Streaming - Imagen)", "Menciones - Empresa", "Mantener"]
    df_final = df.copy()

    st.subheader("📊 Resumen del Proceso")
    col1, col2, col3 = st.columns(3)
    col1.metric("Filas Totales", len(df_final))
    dups_count = (df_final['Estado_Duplicado'] == 'Eliminar').sum()
    col2.metric("Filas Marcadas como Duplicadas", dups_count)
    col3.metric("Filas Únicas", len(df_final) - dups_count)

    excel_data = to_excel_from_df(df_final, final_order)
    st.download_button(
        label="📥 Descargar Archivo Limpio y Mapeado", 
        data=excel_data, 
        file_name=f"Dossier_Limpio_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", 
        mime="application/vnd.openxmlformats-officedocument.sheet"
    )

# ==============================================================================
# INTERFAZ PRINCIPAL DE STREAMLIT
# ==============================================================================

st.title("🚀 Procesador de SOV (Duplicadas) v1.8")
st.markdown("Una herramienta para limpiar, deduplicar y mapear dossieres de noticias.")
st.info("**Instrucciones:**\n\n1. Prepara tu archivo Dossier principal y tu archivo Configuracion.xlsx.\n2. Sube ambos archivos juntos en el área de abajo.\n3. Haz clic en 'Iniciar Proceso'.\n4. La columna **Mantener** te indicará el ID de la noticia original cuando haya duplicados.")

# Información adicional sobre las mejoras
st.success("✅ **Columna Mantener**: Muestra 'Duplicado de: [ID]' para identificar cuál conservar")

with st.expander("Ver estructura requerida para Configuracion.xlsx"):
    st.markdown("- **Regiones**: Columna A (Medio), Columna B (Región).\n- **Internet**: Columna A (Medio Original), Columna B (Medio Mapeado).")

uploaded_files = st.file_uploader(
    "Arrastra y suelta tus archivos aquí (Dossier y Configuracion)", 
    type=["xlsx"], 
    accept_multiple_files=True
)

dossier_file, config_file = None, None

if uploaded_files:
    for file in uploaded_files:
        if 'config' in file.name.lower(): 
            config_file = file
        else: 
            dossier_file = file
    
    if dossier_file: 
        st.success(f"✅ Archivo Dossier cargado: {dossier_file.name}")
    else: 
        st.warning("⚠️ No se ha subido un archivo que parezca ser el Dossier.")
    
    if config_file: 
        st.success(f"✅ Archivo de Configuración cargado: {config_file.name}")
    else: 
        st.warning("⚠️ No se ha subido el archivo Configuracion.xlsx.")

if st.button(
    "▶️ Iniciar Proceso de Limpieza", 
    disabled=not (dossier_file and config_file), 
    type="primary"
):
    run_full_process(dossier_file, config_file)
